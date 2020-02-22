#below code was executed in a jupyter notebook
import openpyxl
import xlrd

wb = openpyxl.load_workbook('Original Data_Tree_CDOC.xlsx')

sheets = wb.sheetnames

sheets

from openpyxl import load_workbook
wb = openpyxl.load_workbook('Original Data_Tree_CDOC.xlsx')
sheets = wb.sheetnames

sheets

sheet = wb['Sheet_Name'] #Cannot supply clients data on public repository. 
                        #Sheet_Name is a placeholder

for i in range(1, 123, 1):
    if sheet.cell(row=i, column=1).value is None:
        print(");") #end of creating a table
        
    elif (sheet.cell(row=i, column=1).fill.fgColor.rgb) == 'FFFFFF00':
        table_name = sheet.cell(row=i, column=1).value
        print ('CREATE TABLE ' + table_name + '(\n')
    
    elif sheet.cell(row=i, column=1).value is None or sheet.cell(row=i+1, column=1).fill.fgColor.rgb == 'FFFFFF00':
        print()
        
    else:
        print(sheet.cell(row=i, column=1).value, ',') #print the comma

        
   
        
       
            
            

